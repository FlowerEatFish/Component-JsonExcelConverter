from openpyxl import load_workbook

class ReadExcel():
    data = {
        'sheet': []
    }

    def __init__(self):
        filePath = self.setPath()
        fileData = self.getFile(filePath)
        self.data = self.getAllSheet(fileData, self.data)
        self.data = self.getAllSheetHeader(fileData, self.data)
        self.data = self.getAllSheetContext(fileData, self.data)
        print('End ReadExcel().')

    def setPath(self):
        print('Please enter the path for excel file.')
        while True:
            inputPath = input()
            if self.isValidFile(inputPath):
                return inputPath

    def isValidFile(self, path):
        try:
            load_workbook(filename = path)
            return True
        except:
            return False

    def getFile(self, path):
        print('Collecting %s ...' % path)
        excelFile = load_workbook(filename = path)
        return excelFile

    def getAllSheet(self, fileContext, data):
        allSheet = fileContext.get_sheet_names()
        for i in allSheet:
            data['sheet'].append(i)
        return data

    def getAllSheetHeader(self, fileContext, data):
        allSheet = data['sheet']
        data['header'] = []

        sheetHeader = []
        for i in range(len(allSheet)):
            targetSheet = fileContext[fileContext.get_sheet_names()[i]]

            sheetHeader = self.getSheetHeader(targetSheet)
            data['header'].append(sheetHeader)
        return data

    def getSheetHeader(self, fileContext):
        sheetHeader = []

        text = ''
        for column in range(26):
            text = fileContext['%s1' % chr(ord('A')+column)].value
            if self.isValidText(text):
                sheetHeader.append(text)
        return sheetHeader

    def getAllSheetContext(self, fileContext, data):
        allSheet = data['sheet']
        data['context'] = []

        allSheetContext = []
        for i in range(len(allSheet)):
            targetSheet = fileContext[fileContext.get_sheet_names()[i]]
            intTargetHeaderLength = len(data['header'][i])

            sheetContext = self.getSheetContext(targetSheet, intTargetHeaderLength)
            data['context'].append(sheetContext)
        return data

    def getSheetContext(self, fileContext, sheetHeaderLength):
        sheetContext = []

        row = 2 # skip row 1 as header.
        while True:
            text = fileContext['A%d' % row].value
            if self.isValidText(text):
                rowContext = []
                for column in range(sheetHeaderLength):
                    text = fileContext['%s%d' % (chr(ord('A')+column), row)].value
                    rowContext.append(text)
                sheetContext.append(rowContext)
                row += 1
            else:
                break
        return sheetContext

    def isValidText(self, text):
        if text != None:
            return True
        return False

# Demo
if __name__ == '__main__':
    result = ReadExcel()
    print(result.data)
